from __future__ import annotations

import ast
import json
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

from .audit import AuditLogger
from .extractors import dump_json, extract_fix_actions, extract_replace_instruction
from .indexer import WikiIndex
from .question_parser import parse_question
from .security import PermissionPolicy, REJECT_MESSAGE


class AnswerEngine:
    def __init__(self, index: WikiIndex, policy: PermissionPolicy, project_root: Path, output_root: Path) -> None:
        self.index = index
        self.policy = policy
        self.project_root = project_root
        self.output_root = output_root
        self.fixed_root = output_root / "fixed"
        self.fixed_root.mkdir(parents=True, exist_ok=True)
        self.audit = AuditLogger(output_root / "audit.jsonl")

    def answer_group(self, question_path: Path, answer_path: Path) -> list[dict[str, object]]:
        payload = json.loads(question_path.read_text(encoding="utf-8"))
        answers = []
        for question in payload:
            answer = self.answer_question(question["title"])
            answers.append(
                {
                    "id": question["id"],
                    "title": question["title"],
                    "level": question["level"],
                    "answer": answer,
                }
            )
        answer_path.parent.mkdir(parents=True, exist_ok=True)
        answer_path.write_text(dump_json(answers), encoding="utf-8")
        self.audit.write("answer_group", question_path=question_path.as_posix(), answer_path=answer_path.as_posix())
        return answers

    def answer_all_groups(self, question_root: Path, output_root: Path) -> list[str]:
        produced: list[str] = []
        for question_path in sorted(question_root.glob("group-*.md")):
            answer_path = output_root / question_path.name.replace(".md", "-answer.md")
            self.answer_group(question_path, answer_path)
            produced.append(answer_path.as_posix())
        self.audit.write("answer_all_groups", count=len(produced))
        return produced

    def answer_question(self, title: str) -> dict[str, object]:
        parsed = parse_question(title, self.index.list_document_paths())
        risk = self.policy.detect_question_risk(parsed.normalized_title, parsed.candidate_paths)
        if risk:
            self.audit.write("question_blocked", title=title, reason=risk["error_msg"])
            return risk

        if parsed.kind == "count_extension" and parsed.extension:
            answer = {parsed.extension: self.index.count_files_by_extension(parsed.extension)}
        elif parsed.kind == "count_supported_extensions":
            answer = self.index.count_supported_extensions()
        elif parsed.kind == "find_path":
            answer = {"datas": parsed.candidate_paths}
        elif parsed.kind == "file_secret_lookup":
            answer = self.lookup_file_secret(parsed.candidate_paths)
        elif parsed.kind == "comment_count":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            answer = {"count": len(self.index.list_comments(rel_path=rel_path))} if rel_path else {"count": 0}
        elif parsed.kind == "assignee_comments":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            rows = self.index.list_comments(rel_path=rel_path, assignee=parsed.assignee) if rel_path else []
            answer = {"datas": [row["text"] for row in rows]}
        elif parsed.kind == "date_comments":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            rows = self.index.list_comments(rel_path=rel_path, due_date=parsed.date_value) if rel_path else []
            answer = {"datas": [row["text"] for row in rows]}
        elif parsed.kind == "global_assignee_comments":
            rows = self.index.list_comments(assignee=parsed.assignee)
            answer = {"datas": [f"{row['rel_path']} | {row['text']}" for row in rows]}
        elif parsed.kind == "global_date_comments":
            rows = self.index.list_comments(due_date=parsed.date_value)
            answer = {"datas": [f"{row['rel_path']} | {row['text']}" for row in rows]}
        elif parsed.kind == "fix_document":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            answer = self.apply_fixes(rel_path) if rel_path else {"datas": []}
        elif parsed.kind == "pivot_chart":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            answer = self.build_pivot_chart(rel_path) if rel_path else {"datas": []}
        elif parsed.kind == "run_document":
            rel_path = self._pick_single_path(parsed.candidate_paths)
            answer = self.run_python_document(rel_path) if rel_path else {"datas": []}
        elif parsed.kind == "run_document_by_keyword" and parsed.keyword:
            rel_path = self._select_keyword_document(parsed.keyword, {".py"})
            answer = self.run_python_document(rel_path) if rel_path else {"datas": []}
        elif parsed.kind == "pivot_chart_by_keyword" and parsed.keyword:
            rel_path = self._select_keyword_document(parsed.keyword, {".xlsx", ".xls"})
            answer = self.build_pivot_chart(rel_path) if rel_path else {"datas": []}
        elif parsed.kind == "business_search" and parsed.keyword:
            answer = {"datas": self.index.search_related_paths(parsed.keyword)}
        elif parsed.kind == "command_lookup" and parsed.keyword:
            answer = {"datas": self._search_command_knowledge(parsed.keyword)}
        elif parsed.kind == "generic_file_lookup":
            answer = {"datas": parsed.candidate_paths}
        elif parsed.keyword:
            answer = {"datas": self.index.search_related_paths(parsed.keyword)}
        else:
            answer = {"datas": []}

        self.audit.write("question_answered", title=title, kind=parsed.kind)
        return answer

    def apply_fixes(self, rel_path: str) -> dict[str, str]:
        absolute_path = self.index.get_document_absolute_path(rel_path)
        extension = self.index.get_document_extension(rel_path)
        if not absolute_path or not extension:
            return {"source": rel_path, "target": ""}
        source_path = Path(absolute_path)
        target_path = self.fixed_root / source_path.name
        comments = [row["text"] for row in self.index.list_comments(rel_path=rel_path)]
        actions = [action for comment in comments for action in extract_fix_actions(comment)]
        if extension in {"py", "java", "js", "md", "html", "xml"}:
            text = source_path.read_text(encoding="utf-8", errors="ignore")
            updated = self._apply_text_actions(text, actions)
            target_path.write_text(updated, encoding="utf-8")
        elif extension == "docx":
            self._apply_docx_actions(source_path, target_path, actions)
        elif extension == "xlsx":
            self._apply_xlsx_actions(source_path, target_path, actions)
        elif extension == "pptx":
            self._apply_pptx_actions(source_path, target_path, actions)
        else:
            shutil.copy2(source_path, target_path)
        report_path = target_path.with_suffix(target_path.suffix + ".fix-report.md")
        report = {
            "source": rel_path,
            "target": f"output/fixed/{source_path.name}",
            "applied_rules": [
                {"kind": action.kind, "source": action.source, "target": action.target, "field_name": action.field_name}
                for action in actions
                if action.kind != "manual"
            ],
            "manual_follow_up": [action.source for action in actions if action.kind == "manual"],
        }
        report_path.write_text(dump_json(report), encoding="utf-8")
        self.audit.write("document_fixed", source=rel_path, target=report["target"])
        return {"source": rel_path, "target": f"output/fixed/{source_path.name}"}

    def build_pivot_chart(self, rel_path: str) -> dict[str, object]:
        absolute_path = self.index.get_document_absolute_path(rel_path)
        if not absolute_path:
            return {"datas": []}
        frame = pd.read_excel(absolute_path)
        categorical = next((column for column in frame.columns if frame[column].dtype == "object"), None)
        numeric = next((column for column in frame.columns if pd.api.types.is_numeric_dtype(frame[column])), None)
        if categorical is None or numeric is None:
            return {"datas": []}
        pivot = frame.pivot_table(index=categorical, values=numeric, aggfunc="sum").sort_values(numeric, ascending=False)
        output_path = self.output_root / f"{Path(rel_path).stem}-pivot.png"
        plt.figure(figsize=(8, 4.5))
        pivot[numeric].plot(kind="bar", title=f"{categorical} by {numeric}")
        plt.tight_layout()
        plt.savefig(output_path, dpi=180)
        plt.close()
        self.audit.write("pivot_generated", source=rel_path, output=f"output/{output_path.name}")
        return {"datas": [f"output/{output_path.name}"]}

    def lookup_file_secret(self, candidate_paths: list[str]) -> dict[str, object]:
        allowed_paths = [path for path in candidate_paths if "02_环境信息" in path.replace("\\", "/")]
        if not allowed_paths:
            return {"datas": []}
        values: list[str] = []
        for rel_path in allowed_paths:
            absolute_path = self.index.get_document_absolute_path(rel_path)
            if not absolute_path:
                continue
            values.extend(self._extract_secret_values(Path(absolute_path)))
        deduped = list(dict.fromkeys(values))
        self.audit.write("file_secret_lookup", paths=allowed_paths, count=len(deduped))
        return {"datas": deduped}

    def run_python_document(self, rel_path: str) -> dict[str, object]:
        absolute_path = self.index.get_document_absolute_path(rel_path)
        if not absolute_path:
            return {"datas": []}
        source = Path(absolute_path).read_text(encoding="utf-8", errors="ignore")
        if not self._is_python_safe(source):
            self.audit.write("python_execution_blocked", source=rel_path)
            return REJECT_MESSAGE
        result = subprocess.run(
            [sys.executable, absolute_path],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
            cwd=str(Path(absolute_path).parent),
        )
        output = result.stdout.strip() or result.stderr.strip()
        self.audit.write("python_executed", source=rel_path, returncode=result.returncode)
        return {"datas": [output]}

    def _is_python_safe(self, source: str) -> bool:
        banned_imports = {"os", "subprocess", "shutil", "pathlib", "socket"}
        banned_calls = {"open", "exec", "eval", "__import__"}
        tree = ast.parse(source)
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    if alias.name.split(".")[0] in banned_imports:
                        return False
            if isinstance(node, ast.ImportFrom):
                if node.module and node.module.split(".")[0] in banned_imports:
                    return False
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id in banned_calls:
                return False
        return True

    def _apply_text_actions(self, text: str, actions: list[object]) -> str:
        updated = text
        for action in actions:
            if getattr(action, "kind", "") == "replace" and action.field_name and action.target is not None:
                old, new = action.field_name, action.target
                updated = updated.replace(old, new)
        add_fields = [action.field_name for action in actions if getattr(action, "kind", "") == "add_field" and action.field_name]
        if add_fields:
            section = "\n".join(f"- 补充字段: {field}" for field in dict.fromkeys(add_fields))
            updated = updated.rstrip() + "\n\n[自动补充字段建议]\n" + section + "\n"
        return updated

    def _apply_docx_actions(self, source_path: Path, target_path: Path, actions: list[object]) -> None:
        with zipfile.ZipFile(source_path) as source_archive:
            files = {name: source_archive.read(name) for name in source_archive.namelist()}
        document_xml = files.get("word/document.xml")
        if document_xml is not None:
            text = document_xml.decode("utf-8", errors="ignore")
            for action in actions:
                if getattr(action, "kind", "") == "replace" and action.field_name and action.target is not None:
                    text = text.replace(action.field_name, action.target)
            add_fields = [action.field_name for action in actions if getattr(action, "kind", "") == "add_field" and action.field_name]
            for field in dict.fromkeys(add_fields):
                insert = (
                    "<w:p><w:r><w:t>"
                    + f"补充字段: {field}"
                    + "</w:t></w:r></w:p>"
                )
                text = text.replace("</w:body>", insert + "</w:body>")
            files["word/document.xml"] = text.encode("utf-8")
        with zipfile.ZipFile(target_path, "w") as target_archive:
            for name, payload in files.items():
                target_archive.writestr(name, payload)

    def _apply_xlsx_actions(self, source_path: Path, target_path: Path, actions: list[object]) -> None:
        workbook = load_workbook(source_path)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        updated = cell.value
                        for action in actions:
                            if getattr(action, "kind", "") == "replace" and action.field_name and action.target is not None:
                                updated = updated.replace(action.field_name, action.target)
                        cell.value = updated
            add_fields = [action.field_name for action in actions if getattr(action, "kind", "") == "add_field" and action.field_name]
            for field in dict.fromkeys(add_fields):
                max_col = sheet.max_column + 1
                sheet.cell(row=1, column=max_col, value=field)
        workbook.save(target_path)

    def _apply_pptx_actions(self, source_path: Path, target_path: Path, actions: list[object]) -> None:
        with zipfile.ZipFile(source_path) as source_archive:
            files = {name: source_archive.read(name) for name in source_archive.namelist()}
        for name, payload in list(files.items()):
            if not name.startswith("ppt/slides/slide") or not name.endswith(".xml"):
                continue
            text = payload.decode("utf-8", errors="ignore")
            for action in actions:
                if getattr(action, "kind", "") == "replace" and action.field_name and action.target is not None:
                    text = text.replace(action.field_name, action.target)
            add_fields = [action.field_name for action in actions if getattr(action, "kind", "") == "add_field" and action.field_name]
            for field in dict.fromkeys(add_fields):
                marker = (
                    '<p:sp><p:txBody><a:p><a:r><a:t>'
                    + f"补充字段: {field}"
                    + "</a:t></a:r></a:p></p:txBody></p:sp>"
                )
                text = text.replace("</p:spTree>", marker + "</p:spTree>")
            files[name] = text.encode("utf-8")
        with zipfile.ZipFile(target_path, "w") as target_archive:
            for name, payload in files.items():
                target_archive.writestr(name, payload)

    def _pick_single_path(self, paths: list[str]) -> str | None:
        return paths[0] if paths else None

    def _select_keyword_document(self, keyword: str, suffixes: set[str]) -> str | None:
        candidates = [keyword]
        candidates.extend(self._extract_search_terms(keyword))
        seen: set[str] = set()
        for candidate in candidates:
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            related_paths = self.index.search_related_paths(candidate)
            for rel_path in related_paths:
                if Path(rel_path).suffix.lower() in suffixes:
                    return rel_path
            parent_dirs = {str(Path(rel_path).parent).replace("\\", "/") for rel_path in related_paths}
            for rel_path in self.index.list_document_paths():
                rel_parent = str(Path(rel_path).parent).replace("\\", "/")
                if rel_parent in parent_dirs and Path(rel_path).suffix.lower() in suffixes:
                    return rel_path
        return None

    def _extract_search_terms(self, text: str) -> list[str]:
        stopwords = {
            "根据",
            "相关",
            "脚本",
            "执行结果",
            "给出",
            "输出",
            "运行结果",
            "文件",
            "代码",
            "结果",
            "如何",
            "怎么",
            "在",
            "控制台",
            "连接",
        }
        tokens = re.findall(r"[A-Za-z0-9_.+-]+|[\u4e00-\u9fff]{2,}", text)
        results: list[str] = []
        for token in tokens:
            if token in stopwords:
                continue
            results.append(token)
            if "高斯数据库" in token and "gauss" not in results:
                results.append("gauss")
            if len(token) > 4:
                for splitter in ("如何", "怎么", "控制台", "连接", "相关", "脚本", "执行结果", "运行结果"):
                    token = token.replace(splitter, " ")
                parts = [part.strip() for part in token.split() if part.strip() and part.strip() not in stopwords]
                for part in parts:
                    if part not in results:
                        results.append(part)
        return results

    def _search_command_knowledge(self, text: str) -> list[str]:
        candidates = [text]
        candidates.extend(self._extract_search_terms(text))
        results: list[str] = []
        seen: set[str] = set()
        for candidate in candidates:
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            for snippet in self.index.search_snippets(candidate):
                if snippet not in results:
                    results.append(snippet)
            if results:
                break
        return results

    def _extract_secret_values(self, path: Path) -> list[str]:
        suffix = path.suffix.lower()
        text = path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".xml":
            values = self._extract_xml_secret_values(text)
            if values:
                return values
        return self._extract_text_secret_values(text)

    def _extract_xml_secret_values(self, text: str) -> list[str]:
        try:
            root = ET.fromstring(text)
        except ET.ParseError:
            return []
        values: list[str] = []
        for node in root.iter():
            tag = node.tag.split("}")[-1].lower()
            if tag in {"password", "passwd", "secret", "token", "key"} and (node.text or "").strip():
                values.append(node.text.strip())
            for attr_name, attr_value in node.attrib.items():
                lowered = attr_name.lower()
                if lowered in {"password", "passwd", "secret", "token", "key"} and attr_value.strip():
                    values.append(attr_value.strip())
        return values

    def _extract_text_secret_values(self, text: str) -> list[str]:
        patterns = (
            re.compile(r"password\s*[:=]\s*['\"]?([^'\"\n\r]+)", re.IGNORECASE),
            re.compile(r"passwd\s*[:=]\s*['\"]?([^'\"\n\r]+)", re.IGNORECASE),
            re.compile(r"secret\s*[:=]\s*['\"]?([^'\"\n\r]+)", re.IGNORECASE),
            re.compile(r"token\s*[:=]\s*['\"]?([^'\"\n\r]+)", re.IGNORECASE),
        )
        values: list[str] = []
        for pattern in patterns:
            for match in pattern.finditer(text):
                values.append(match.group(1).strip())
        return values
