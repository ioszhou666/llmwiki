from __future__ import annotations

import html.parser
import json
import re
import shutil
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from .constants import SUPPORTED_EXTENSIONS, TEXT_EXTENSIONS
from .models import CommentRecord, DocumentRecord


OOXML_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
}
STRUCTURED_TODO_RE = re.compile(
    r"todo\s*[:：]\s*(?P<todo>.+?)\s*[，,]\s*to\s*[:：]\s*(?P<to>.+?)\s*[，,]\s*end_date\s*[:：]\s*(?P<end_date>\d{8})",
    re.IGNORECASE,
)
REPLACE_PATTERNS = (
    re.compile(r"把(?P<old>.+?)改成(?P<new>.+)"),
    re.compile(r"将(?P<old>.+?)改为(?P<new>.+)"),
    re.compile(r"replace\s+(?P<old>.+?)\s+with\s+(?P<new>.+)", re.IGNORECASE),
)
ADD_FIELD_PATTERNS = (
    re.compile(r"补充(?P<field>.+?)字段"),
    re.compile(r"新增(?P<field>.+?)字段"),
    re.compile(r"add\s+(?P<field>.+?)\s+field", re.IGNORECASE),
)


@dataclass(slots=True)
class FixAction:
    kind: str
    source: str
    target: str | None = None
    field_name: str | None = None


class _HTMLTextParser(html.parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        text = data.strip()
        if text:
            self.parts.append(text)

    def get_text(self) -> str:
        return "\n".join(self.parts)


def discover_documents(docs_root: Path) -> list[Path]:
    return sorted(path for path in docs_root.rglob("*") if path.is_file())


def read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "gbk", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="ignore")


def extract_document(path: Path, project_root: Path) -> DocumentRecord:
    extension = path.suffix.lstrip(".").lower()
    relative_path = path.relative_to(project_root).as_posix()
    if extension == "docx":
        content, comments, metadata = extract_docx(path)
    elif extension == "pptx":
        content, comments, metadata = extract_pptx(path)
    elif extension == "xlsx":
        content, comments, metadata = extract_xlsx(path)
    elif extension in TEXT_EXTENSIONS:
        content, comments, metadata = extract_text_document(path, extension)
    elif extension in {"doc", "ppt", "xls"}:
        content, comments, metadata = extract_legacy_document(path, extension)
    else:
        content = read_text_with_fallback(path)
        comments = []
        metadata = {}
    return DocumentRecord(
        absolute_path=str(path.resolve()),
        relative_path=relative_path,
        extension=extension,
        content=content,
        comments=comments,
        metadata={"size": path.stat().st_size, **metadata},
    )


def extract_text_document(path: Path, extension: str) -> tuple[str, list[CommentRecord], dict[str, object]]:
    raw_text = read_text_with_fallback(path)
    if extension == "html":
        parser = _HTMLTextParser()
        parser.feed(raw_text)
        content = parser.get_text()
    elif extension == "xml":
        try:
            root = ET.fromstring(raw_text)
            content = "\n".join(part.strip() for part in root.itertext() if part.strip())
        except ET.ParseError:
            content = raw_text
    else:
        content = raw_text
    comments = extract_todos_from_text(raw_text, str(path), extension)
    return content, comments, {"extraction_method": "native_text"}


def extract_docx(path: Path) -> tuple[str, list[CommentRecord], dict[str, object]]:
    comments: list[CommentRecord] = []
    with zipfile.ZipFile(path) as archive:
        document_xml = archive.read("word/document.xml")
        root = ET.fromstring(document_xml)
        content = "\n".join(part.strip() for part in root.itertext() if part and part.strip())
        if "word/comments.xml" in archive.namelist():
            comments_root = ET.fromstring(archive.read("word/comments.xml"))
            for node in comments_root.findall("w:comment", OOXML_NS):
                text = "\n".join(part.strip() for part in node.itertext() if part and part.strip())
                parsed = parse_structured_comment(text)
                comments.append(
                    CommentRecord(
                        path=str(path),
                        file_type="docx",
                        text=text,
                        kind="office_comment",
                        assignee=parsed.get("to"),
                        due_date=parsed.get("end_date"),
                        author=node.attrib.get(f"{{{OOXML_NS['w']}}}author"),
                        created_at=node.attrib.get(f"{{{OOXML_NS['w']}}}date"),
                        location=node.attrib.get(f"{{{OOXML_NS['w']}}}id"),
                        structured=bool(parsed),
                        todo_text=parsed.get("todo"),
                    )
                )
    return content, comments, {"extraction_method": "native_docx"}


def extract_pptx(path: Path) -> tuple[str, list[CommentRecord], dict[str, object]]:
    texts: list[str] = []
    comments: list[CommentRecord] = []
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            if name.startswith("ppt/slides/slide") and name.endswith(".xml"):
                slide_root = ET.fromstring(archive.read(name))
                texts.extend(part.strip() for part in slide_root.itertext() if part and part.strip())
        authors: dict[str, str] = {}
        if "ppt/commentAuthors.xml" in archive.namelist():
            author_root = ET.fromstring(archive.read("ppt/commentAuthors.xml"))
            for node in author_root.findall("p:cmAuthor", OOXML_NS):
                authors[node.attrib.get("id", "")] = node.attrib.get("name", "")
        for name in sorted(archive.namelist()):
            if name.startswith("ppt/comments/comment") and name.endswith(".xml"):
                comment_root = ET.fromstring(archive.read(name))
                for node in comment_root.findall("p:cm", OOXML_NS):
                    text = "\n".join(part.strip() for part in node.itertext() if part and part.strip())
                    parsed = parse_structured_comment(text)
                    comments.append(
                        CommentRecord(
                            path=str(path),
                            file_type="pptx",
                            text=text,
                            kind="office_comment",
                            assignee=parsed.get("to"),
                            due_date=parsed.get("end_date"),
                            author=authors.get(node.attrib.get("authorId", ""), node.attrib.get("authorId")),
                            created_at=node.attrib.get("dt"),
                            location=name,
                            structured=bool(parsed),
                            todo_text=parsed.get("todo"),
                        )
                    )
    return "\n".join(texts), comments, {"extraction_method": "native_pptx"}


def extract_xlsx(path: Path) -> tuple[str, list[CommentRecord], dict[str, object]]:
    workbook = load_workbook(path, data_only=False)
    texts: list[str] = []
    comments: list[CommentRecord] = []
    for sheet in workbook.worksheets:
        texts.append(sheet.title)
        for row in sheet.iter_rows():
            values: list[str] = []
            for cell in row:
                if cell.value not in (None, ""):
                    values.append(str(cell.value))
                if cell.comment:
                    parsed = parse_structured_comment(cell.comment.text)
                    comments.append(
                        CommentRecord(
                            path=str(path),
                            file_type="xlsx",
                            text=cell.comment.text,
                            kind="office_comment",
                            assignee=parsed.get("to"),
                            due_date=parsed.get("end_date"),
                            author=cell.comment.author,
                            location=f"{sheet.title}!{cell.coordinate}",
                            structured=bool(parsed),
                            todo_text=parsed.get("todo"),
                        )
                    )
            if values:
                texts.append(" | ".join(values))
    return "\n".join(texts), comments, {"extraction_method": "native_xlsx"}


def extract_legacy_document(path: Path, extension: str) -> tuple[str, list[CommentRecord], dict[str, object]]:
    converted = convert_legacy_office(path, extension)
    if converted is not None:
        try:
            if extension == "doc":
                content, comments, _ = extract_docx(converted)
                return content, comments, {"extraction_method": "libreoffice_convert"}
            if extension == "ppt":
                content, comments, _ = extract_pptx(converted)
                return content, comments, {"extraction_method": "libreoffice_convert"}
            if extension == "xls":
                content, comments, _ = extract_xlsx(converted)
                return content, comments, {"extraction_method": "libreoffice_convert"}
        finally:
            shutil.rmtree(converted.parent, ignore_errors=True)
    tika_jar = locate_tika_jar()
    if tika_jar is not None:
        result = subprocess.run(
            ["java", "-jar", str(tika_jar), "--text", str(path)],
            capture_output=True,
            text=True,
            timeout=30,
            check=False,
        )
        if result.returncode == 0:
            return result.stdout, [], {"extraction_method": "tika_text"}
    fallback = read_text_with_fallback(path)
    note = f"[legacy-{extension}-fallback] 未检测到可用转换器，当前内容可能不完整。"
    return f"{note}\n{fallback}", [], {"extraction_method": "plain_fallback"}


def locate_tika_jar() -> Path | None:
    candidates = [
        Path.cwd() / "tools" / "tika-app.jar",
        Path.cwd() / "work" / "llm_wiki" / "tools" / "tika-app.jar",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def convert_legacy_office(path: Path, extension: str) -> Path | None:
    soffice = shutil.which("soffice")
    if not soffice:
        return None
    suffix_map = {"doc": "docx", "ppt": "pptx", "xls": "xlsx"}
    target_extension = suffix_map.get(extension)
    if not target_extension:
        return None
    temp_dir_path = Path(tempfile.mkdtemp(prefix="llm_wiki_legacy_"))
    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            target_extension,
            "--outdir",
            str(temp_dir_path),
            str(path),
        ],
        capture_output=True,
        text=True,
        timeout=60,
        check=False,
    )
    if result.returncode != 0:
        shutil.rmtree(temp_dir_path, ignore_errors=True)
        return None
    converted_path = temp_dir_path / f"{path.stem}.{target_extension}"
    if not converted_path.exists():
        shutil.rmtree(temp_dir_path, ignore_errors=True)
        return None
    return converted_path


def extract_todos_from_text(raw_text: str, path_text: str, extension: str) -> list[CommentRecord]:
    comments: list[CommentRecord] = []
    lines = raw_text.splitlines()
    for line_number, line in enumerate(lines, start=1):
        if "todo" not in line.lower():
            continue
        parsed = parse_structured_comment(line)
        comments.append(
            CommentRecord(
                path=path_text,
                file_type=extension,
                text=line.strip(),
                kind="code_todo",
                assignee=parsed.get("to"),
                due_date=parsed.get("end_date"),
                location=f"line:{line_number}",
                structured=bool(parsed),
                todo_text=parsed.get("todo"),
            )
        )
    for block in re.findall(r"/\*(.*?)\*/", raw_text, flags=re.DOTALL):
        stripped = block.strip()
        if stripped:
            comments.append(CommentRecord(path=path_text, file_type=extension, text=stripped, kind="code_comment"))
    for block in re.findall(r"<!--(.*?)-->", raw_text, flags=re.DOTALL):
        stripped = block.strip()
        if stripped:
            comments.append(CommentRecord(path=path_text, file_type=extension, text=stripped, kind="code_comment"))
    return comments


def parse_structured_comment(text: str) -> dict[str, str]:
    match = STRUCTURED_TODO_RE.search(text)
    if not match:
        return {}
    return {key: value.strip() for key, value in match.groupdict().items()}


def extract_fix_actions(text: str) -> list[FixAction]:
    structured = parse_structured_comment(text)
    source_text = structured.get("todo", text)
    clean = re.sub(r"\s+", " ", source_text).strip()
    actions: list[FixAction] = []
    for pattern in REPLACE_PATTERNS:
        match = pattern.search(clean)
        if match:
            actions.append(FixAction(kind="replace", source=clean, target=match.group("new").strip(" ，,。.;；"), field_name=match.group("old").strip(" ，,。.;；")))
            return actions
    for pattern in ADD_FIELD_PATTERNS:
        match = pattern.search(clean)
        if match:
            actions.append(FixAction(kind="add_field", source=clean, field_name=match.group("field").strip(" ，,。.;；")))
            return actions
    if clean:
        actions.append(FixAction(kind="manual", source=clean))
    return actions


def extract_replace_instruction(text: str) -> tuple[str, str] | None:
    for action in extract_fix_actions(text):
        if action.kind == "replace" and action.field_name and action.target is not None:
            return action.field_name, action.target
    return None


def extract_add_field_instruction(text: str) -> str | None:
    for action in extract_fix_actions(text):
        if action.kind == "add_field" and action.field_name:
            return action.field_name
    return None


def dump_json(data: object) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)
