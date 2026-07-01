from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment

from llm_wiki.answerer import AnswerEngine
from llm_wiki.demo_workspace import build_sample_workspace, write_text
from llm_wiki.extractors import extract_document, extract_fix_actions
from llm_wiki.indexer import WikiIndex
from llm_wiki.security import PermissionPolicy


def test_extract_fix_actions_supports_replace_and_add_field() -> None:
    replace_actions = extract_fix_actions("todo: 把旧标题改成新标题, to: 张三, end_date: 20261231")
    assert replace_actions[0].kind == "replace"
    assert replace_actions[0].field_name == "旧标题"
    assert replace_actions[0].target == "新标题"

    add_actions = extract_fix_actions("todo: 补充产品报价字段, to: 张三, end_date: 20261231")
    assert add_actions[0].kind == "add_field"
    assert add_actions[0].field_name == "产品报价"


def test_legacy_document_returns_fallback_metadata(tmp_path: Path) -> None:
    legacy_path = tmp_path / "docs" / "07_其他" / "legacy.doc"
    write_text(legacy_path, "legacy plain fallback")
    record = extract_document(legacy_path, tmp_path)
    assert record.metadata["extraction_method"] == "plain_fallback"
    assert "legacy plain fallback" in record.content


def test_add_field_fix_updates_text_and_xlsx(tmp_path: Path) -> None:
    build_sample_workspace(tmp_path)
    write_text(
        tmp_path / "docs" / "05_需求设计" / "field_note.md",
        "<!-- todo: 补充产品报价字段, to: 张三, end_date: 20261231 -->\n当前文档说明",
    )
    db_path = tmp_path / "output" / "wiki.db"
    index = WikiIndex(db_path)
    try:
        index.index_documents(tmp_path / "docs", tmp_path)
        policy = PermissionPolicy.from_file(tmp_path / "Permission.json")
        engine = AnswerEngine(index=index, policy=policy, project_root=tmp_path, output_root=tmp_path / "output")

        text_fix = engine.apply_fixes("docs/05_需求设计/field_note.md")
        fixed_text = (tmp_path / "output" / "fixed" / "field_note.md").read_text(encoding="utf-8")
        assert text_fix["target"] == "output/fixed/field_note.md"
        assert "补充字段: 产品报价" in fixed_text

        workbook_path = tmp_path / "docs" / "06_日常办公" / "销售统计.xlsx"
        workbook = load_workbook(workbook_path)
        sheet = workbook["销售"]
        sheet["B1"].comment = Comment("todo: 补充产品报价字段, to: 张三, end_date: 20261231", "赵六")
        workbook.save(workbook_path)
        index.index_documents(tmp_path / "docs", tmp_path)

        xlsx_fix = engine.apply_fixes("docs/06_日常办公/销售统计.xlsx")
        fixed_book = load_workbook(tmp_path / "output" / "fixed" / "销售统计.xlsx")
        fixed_sheet = fixed_book["销售"]
        headers = [fixed_sheet.cell(row=1, column=idx).value for idx in range(1, fixed_sheet.max_column + 1)]
        assert xlsx_fix["target"] == "output/fixed/销售统计.xlsx"
        assert "产品报价" in headers
    finally:
        index.close()
